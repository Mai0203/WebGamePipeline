"""读取输入文件并统一规范成 `InputRecord` 列表。"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from .jsonl_utils import iter_jsonl


@dataclass(slots=True)
class InputRecord:
    """单条输入样本的标准结构。"""
    data_id: str
    query: str
    payload: dict[str, Any]


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    """读取 CSV 文件。"""
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_openpyxl_rows(path: Path) -> list[dict[str, Any]]:
    """使用 openpyxl 读取常见的 xlsx 文件。"""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - 可选运行时依赖
        raise RuntimeError(
            "Reading .xlsx files requires openpyxl. Install it with `pip install openpyxl`."
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    row_iter = worksheet.iter_rows(values_only=True)
    header = next(row_iter, None)
    if not header:
        workbook.close()
        return []

    header_cells = [str(cell).strip() if cell is not None else "" for cell in header]
    rows: list[dict[str, Any]] = []

    for row in row_iter:
        record: dict[str, Any] = {}
        for index, key in enumerate(header_cells):
            if not key:
                continue
            record[key] = row[index] if index < len(row) else None
        rows.append(record)

    workbook.close()
    return rows


def read_pandas_excel_rows(path: Path) -> list[dict[str, Any]]:
    """使用 pandas 兜底读取其他 Excel 格式。"""
    try:
        import pandas as pd
    except ImportError as exc:  # pragma: no cover - 可选运行时依赖
        raise RuntimeError(
            "Reading this Excel format requires pandas. Install it with `pip install pandas`."
        ) from exc

    frame = pd.read_excel(path)
    return frame.to_dict("records")


def read_excel_rows(path: Path) -> list[dict[str, Any]]:
    """根据后缀选择合适的 Excel 读取方式。"""
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return read_openpyxl_rows(path)
    return read_pandas_excel_rows(path)


def normalize_cell(value: Any) -> str:
    """把单元格值规整成去空白后的字符串。"""
    if value is None:
        return ""
    return str(value).strip()


def validate_record(
    payload: dict[str, Any],
    *,
    data_id_field: str,
    query_field: str,
    data_id_prefix: str,
) -> InputRecord | None:
    """校验并抽取单条记录中的 `data_id` 和 `query`。"""
    if data_id_field not in payload or query_field not in payload:
        missing = []
        if data_id_field not in payload:
            missing.append(data_id_field)
        if query_field not in payload:
            missing.append(query_field)
        raise ValueError(f"Input record missing required fields: {', '.join(missing)}")

    data_id = normalize_cell(payload.get(data_id_field))
    query = normalize_cell(payload.get(query_field))
    if not data_id or not query:
        return None
    if data_id_prefix and not data_id.startswith(data_id_prefix):
        return None

    normalized_payload = dict(payload)
    normalized_payload["data_id"] = data_id
    normalized_payload["query"] = query
    return InputRecord(data_id=data_id, query=query, payload=normalized_payload)


def load_input_records(config) -> list[InputRecord]:
    """按配置读取输入文件，并做过滤、截断和字段标准化。"""
    path = config.path
    if not path.exists():
        raise FileNotFoundError(f"Input file does not exist: {path}")
    if config.limit == 0:
        return []

    suffix = path.suffix.lower()
    if suffix == ".jsonl":
        raw_rows = list(iter_jsonl(path))
    elif suffix == ".csv":
        raw_rows = read_csv_rows(path)
    elif suffix in {".xlsx", ".xls", ".xlsm", ".xltx", ".xltm"}:
        raw_rows = read_excel_rows(path)
    else:
        raise ValueError(
            f"Unsupported input file format: {path}. "
            "Supported: .jsonl, .csv, .xlsx, .xls"
        )

    records: list[InputRecord] = []
    for payload in raw_rows:
        record = validate_record(
            payload,
            data_id_field=config.data_id_field,
            query_field=config.query_field,
            data_id_prefix=config.data_id_prefix,
        )
        if record is None:
            continue
        records.append(record)
        if config.limit is not None and len(records) >= config.limit:
            break

    return records
